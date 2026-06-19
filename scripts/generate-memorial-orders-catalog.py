#!/usr/bin/env python3
"""Generate src/data/memorial-orders-catalog.json from scripts/tavozun-template.xlsx."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "scripts" / "tavozun-template.xlsx"
OUTPUT = ROOT / "src" / "data" / "memorial-orders-catalog.json"

SKIP_OPS = {
    "Ташкилот, корҳона",
    "Сарчашмаи маблағгузорӣ",
    "Рамзи ченак:      сомонӣ",
    "2",
    "Номгӯи амалиёт",
    "барои",
}


def norm_acct(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not re.match(r"^\d", text):
        return None
    parts = re.split(r"[.\s]+", text)
    if len(parts) < 2:
        return None
    return " ".join(parts)


def parse_number(sheet_name: str) -> str:
    match = re.search(r"№\s*(\d+[А-Яа-яA-Za-z]?)", sheet_name)
    return match.group(1) if match else sheet_name


def main() -> None:
    wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
    orders = []

    for name in wb.sheetnames:
        if not name.startswith("М.О."):
            continue

        sheet = wb[name]
        number = parse_number(name)
        subtitle = ""
        for row in range(1, 6):
            for col in range(1, 6):
                value = sheet.cell(row, col).value
                if isinstance(value, str):
                    text = value.strip()
                    if text and "Ордери мемориалии" not in text and "№" not in text and len(text) > 8:
                        subtitle = text
                        break
            if subtitle:
                break

        header_row = None
        for row in range(1, 25):
            values = [str(sheet.cell(row, col).value or "") for col in range(1, 9)]
            if any("Дебет" in value for value in values) and any(
                "Кредит" in value for value in values
            ):
                header_row = row
                break

        seen: set[tuple[str, str, str]] = set()
        operations = []
        if header_row:
            for row in range(header_row + 2, sheet.max_row + 1):
                op = sheet.cell(row, 2).value
                if not op:
                    continue
                op_text = str(op).strip()
                if op_text in SKIP_OPS or op_text.isdigit():
                    continue
                debit = norm_acct(sheet.cell(row, 4).value)
                credit = norm_acct(sheet.cell(row, 5).value)
                if not debit or not credit:
                    continue
                key = (op_text.lower(), debit, credit)
                if key in seen:
                    continue
                seen.add(key)
                basis = sheet.cell(row, 7).value
                operations.append(
                    {
                        "id": f"mo{number}-op{len(operations) + 1}",
                        "name": op_text,
                        "debitAccount": debit,
                        "creditAccount": credit,
                        "basisHint": str(basis).strip() if basis else None,
                    }
                )

        order_id = "mo-" + number.lower().replace("б", "b")
        orders.append(
            {
                "id": order_id,
                "number": number,
                "sheetName": name,
                "title": subtitle or name,
                "operations": operations,
            }
        )

    OUTPUT.write_text(json.dumps(orders, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(orders)} memorial orders to {OUTPUT}")


if __name__ == "__main__":
    main()
